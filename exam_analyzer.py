import re
import pandas as pd
from collections import defaultdict
import PyPDF2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime

class ExamResultAnalyzer:
    def __init__(self):
        self.departments = {}
        self.courses = {}
        self.student_data = []
        
    def parse_pdf(self, pdf_path):
        """Parse the exam result PDF and extract student data"""
        try:
            with open(pdf_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                text = ""
                for page in reader.pages:
                    text += page.extract_text()
            
            return self._extract_data_from_text(text)
        except Exception as e:
            print(f"Error reading PDF: {e}")
            return None
    
    def _extract_data_from_text(self, text):
        """Extract structured data from PDF text"""
        lines = text.split('\n')
        current_dept = None
        current_courses = []
        
        for i, line in enumerate(lines):
            # Detect department
            if 'ENGINEERING[Full Time]' in line or 'ENGINEERING & COMMUNICATION' in line:
                current_dept = line.split('[')[0].strip()
                current_courses = []
                
            # Extract course codes and names
            elif line.startswith('GYMAT') or line.startswith('GZPHT') or line.startswith('GCEST') or \
                 line.startswith('UCEST') or line.startswith('GBPHT') or line.startswith('GMEST') or \
                 line.startswith('GXEST') or line.startswith('GAMAT') or line.startswith('GXCYT'):
                course_code = line.split()[0]
                course_name = ' '.join(line.split()[1:])
                current_courses.append({'code': course_code, 'name': course_name})
            
            # Extract student results
            elif line.startswith('AIK24'):
                reg_no = line.split()[0]
                results_str = ' '.join(line.split()[1:])
                
                # Parse individual course results
                results = self._parse_student_results(results_str)
                
                self.student_data.append({
                    'register_no': reg_no,
                    'department': current_dept,
                    'courses': current_courses.copy(),
                    'results': results
                })
        
        return self.student_data
    
    def _parse_student_results(self, results_str):
        """Parse student results string into structured format"""
        results = []
        # Match pattern: COURSECODE(GRADE)
        pattern = r'([A-Z0-9]+)\(([A-Z+\-]+|Absent|F)\)'
        matches = re.findall(pattern, results_str)
        
        for course_code, grade in matches:
            results.append({
                'course_code': course_code,
                'grade': grade
            })
        
        return results
    
    def analyze_performance(self):
        """Analyze overall performance and generate statistics"""
        dept_stats = defaultdict(lambda: {
            'total_students': 0,
            'students_with_arrears': 0,
            'total_arrears': 0,
            'course_failures': defaultdict(int),
            'course_absences': defaultdict(int),
            'students_passed_all': 0
        })
        
        for student in self.student_data:
            dept = student['department']
            if dept:
                stats = dept_stats[dept]
                stats['total_students'] += 1
                
                failed_courses = []
                absent_courses = []
                
                for result in student['results']:
                    grade = result['grade']
                    course = result['course_code']
                    
                    if grade == 'F':
                        failed_courses.append(course)
                        stats['course_failures'][course] += 1
                    elif grade == 'Absent':
                        absent_courses.append(course)
                        stats['course_absences'][course] += 1
                
                arrear_count = len(failed_courses) + len(absent_courses)
                if arrear_count > 0:
                    stats['students_with_arrears'] += 1
                    stats['total_arrears'] += arrear_count
                else:
                    stats['students_passed_all'] += 1
        
        return dict(dept_stats)
    
    def generate_excel_report(self, output_path='exam_analysis.xlsx'):
        """Generate comprehensive Excel report"""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_department_sheets(wb)
        self._create_student_details_sheet(wb)
        self._create_arrear_analysis_sheet(wb)
        
        wb.save(output_path)
        print(f"Excel report generated: {output_path}")
        return output_path
    
    def _create_summary_sheet(self, wb):
        """Create overall summary sheet"""
        ws = wb.create_sheet("Overall Summary", 0)
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws['A1'] = "APJ ABDUL KALAM TECHNOLOGICAL UNIVERSITY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = "Exam Performance Analysis Report"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f"Generated on: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}"
        
        # Department-wise summary
        stats = self.analyze_performance()
        
        row = 5
        ws[f'A{row}'] = "Department"
        ws[f'B{row}'] = "Total Students"
        ws[f'C{row}'] = "Passed All"
        ws[f'D{row}'] = "With Arrears"
        ws[f'E{row}'] = "Total Arrears"
        ws[f'F{row}'] = "Pass %"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].fill = header_fill
            ws[f'{col}{row}'].font = header_font
        
        row += 1
        for dept, data in stats.items():
            ws[f'A{row}'] = dept
            ws[f'B{row}'] = data['total_students']
            ws[f'C{row}'] = data['students_passed_all']
            ws[f'D{row}'] = data['students_with_arrears']
            ws[f'E{row}'] = data['total_arrears']
            pass_pct = (data['students_passed_all'] / data['total_students'] * 100) if data['total_students'] > 0 else 0
            ws[f'F{row}'] = f"{pass_pct:.2f}%"
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
    
    def _create_department_sheets(self, wb):
        """Create individual sheets for each department"""
        dept_data = defaultdict(list)
        
        for student in self.student_data:
            if student['department']:
                dept_data[student['department']].append(student)
        
        for dept, students in dept_data.items():
            # Truncate sheet name if too long
            sheet_name = dept[:31] if len(dept) > 31 else dept
            ws = wb.create_sheet(sheet_name)
            
            # Headers
            ws['A1'] = "Register No"
            ws['B1'] = "Total Arrears"
            ws['C1'] = "Failed Courses"
            ws['D1'] = "Absent Courses"
            ws['E1'] = "Status"
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}1'].fill = header_fill
                ws[f'{col}1'].font = header_font
            
            row = 2
            for student in students:
                failed = [r['course_code'] for r in student['results'] if r['grade'] == 'F']
                absent = [r['course_code'] for r in student['results'] if r['grade'] == 'Absent']
                
                ws[f'A{row}'] = student['register_no']
                ws[f'B{row}'] = len(failed) + len(absent)
                ws[f'C{row}'] = ', '.join(failed) if failed else '-'
                ws[f'D{row}'] = ', '.join(absent) if absent else '-'
                ws[f'E{row}'] = "CLEAR" if (len(failed) + len(absent)) == 0 else "ARREAR"
                
                # Color coding
                if (len(failed) + len(absent)) == 0:
                    ws[f'E{row}'].fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    ws[f'E{row}'].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row += 1
            
            # Adjust columns
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 12
    
    def _create_student_details_sheet(self, wb):
        """Create detailed student-wise sheet"""
        ws = wb.create_sheet("Student Details")
        
        ws['A1'] = "Register No"
        ws['B1'] = "Department"
        ws['C1'] = "Course Code"
        ws['D1'] = "Grade"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}1'].fill = header_fill
            ws[f'{col}1'].font = header_font
        
        row = 2
        for student in self.student_data:
            for result in student['results']:
                ws[f'A{row}'] = student['register_no']
                ws[f'B{row}'] = student['department']
                ws[f'C{row}'] = result['course_code']
                ws[f'D{row}'] = result['grade']
                row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
    
    def _create_arrear_analysis_sheet(self, wb):
        """Create course-wise arrear analysis"""
        ws = wb.create_sheet("Course Analysis")
        
        # Collect course-wise failures
        course_failures = defaultdict(int)
        course_absences = defaultdict(int)
        
        for student in self.student_data:
            for result in student['results']:
                if result['grade'] == 'F':
                    course_failures[result['course_code']] += 1
                elif result['grade'] == 'Absent':
                    course_absences[result['course_code']] += 1
        
        ws['A1'] = "Course Code"
        ws['B1'] = "Failures"
        ws['C1'] = "Absences"
        ws['D1'] = "Total Issues"
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}1'].fill = header_fill
            ws[f'{col}1'].font = header_font
        
        # Get all unique courses
        all_courses = set(list(course_failures.keys()) + list(course_absences.keys()))
        
        row = 2
        for course in sorted(all_courses):
            failures = course_failures[course]
            absences = course_absences[course]
            
            ws[f'A{row}'] = course
            ws[f'B{row}'] = failures
            ws[f'C{row}'] = absences
            ws[f'D{row}'] = failures + absences
            row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15


# Usage Example
if __name__ == "__main__":
    print("=== Exam Result Analyzer ===\n")
    
    analyzer = ExamResultAnalyzer()
    
    # Parse the PDF
    pdf_path = "C:\Users\alexj\Downloads\Copy of result_AIK_S4_APRIL_2025.xlsx"  # Your PDF file path
    print(f"Parsing PDF: {pdf_path}")
    analyzer.parse_pdf(pdf_path)
    
    print(f"Processed {len(analyzer.student_data)} student records\n")
    
    # Generate Excel report
    output_file = 'r"C:\Users\YourName\Documents\exam_analysis.xlsx"'
    analyzer.generate_excel_report(output_file)
    
    # Print summary statistics
    stats = analyzer.analyze_performance()
    print("\n=== Summary Statistics ===")
    for dept, data in stats.items():
        print(f"\n{dept}:")
        print(f"  Total Students: {data['total_students']}")
        print(f"  Passed All: {data['students_passed_all']}")
        print(f"  With Arrears: {data['students_with_arrears']}")
        print(f"  Total Arrears: {data['total_arrears']}")