from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime

print("="*60)
print("PHASE 4: CREATING EXCEL REPORTS")
print("="*60)

# ========================================
# LESSON 1: Excel Structure Basics
# ========================================
print("\nLESSON 1: Understanding Excel Structure")
print("-"*60)
print("""
Excel Hierarchy:
    Workbook (the .xlsx file)
      └── Worksheet (Sheet1, Sheet2, etc.)
            └── Cell (A1, B2, C3, etc.)

Cells are referenced by:
    - Letter + Number: A1, B2, Z99
    - Row/Column index: row=1, column=1 is A1
    
Python equivalents:
    Workbook = Excel file
    Worksheet = Tab/Sheet
    Cell = Single box
""")

# ========================================
# LESSON 2: Creating Your First Excel
# ========================================
print("\n" + "="*60)
print("LESSON 2: Create Your First Excel File")
print("-"*60)

# Step 1: Create a new workbook
wb = Workbook()

# Step 2: Get the active sheet (default sheet)
ws = wb.active
ws.title = "Student Results"  # Rename the sheet

# Step 3: Write data to cells
ws['A1'] = "Register No"
ws['B1'] = "Name"
ws['C1'] = "Score"

# Step 4: Add student data
ws['A2'] = "AIK24CS001"
ws['B2'] = "John"
ws['C2'] = 85

ws['A3'] = "AIK24CS002"
ws['B3'] = "Jane"
ws['C3'] = 92

# Step 5: Save the file
wb.save('my_first_excel.xlsx')
print("✓ Created 'my_first_excel.xlsx'")
print("  Open it and see your data!")

print("\n" + "="*60)
print("LESSON 3: Better Ways to Write Data")
print("-"*60)

# Method 1: Using row/column numbers
wb = Workbook()
ws = wb.active

ws.cell(row=1, column=1, value="Method 1")
ws.cell(row=1, column=2, value="Using cell()")
print("Method 1: ws.cell(row=1, column=1, value='Data')")

# Method 2: Using append (adds new row)
ws.append(["Method 2", "Using append()"])
print("Method 2: ws.append(['Data1', 'Data2'])")

# Method 3: Loop through data
students = [
    ["AIK24CS001", "John", 85],
    ["AIK24CS002", "Jane", 92],
    ["AIK24CS003", "Bob", 78],
]

ws.append(["Register", "Name", "Score"])  # Header
for student in students:
    ws.append(student)

print("Method 3: Loop through list of lists")

wb.save('writing_methods.xlsx')
print("✓ Created 'writing_methods.xlsx'")

print("\n" + "="*60)
print("LESSON 4: Making It Look Professional")
print("-"*60)

wb = Workbook()
ws = wb.active
ws.title = "Styled Report"

# Headers
headers = ["Register No", "Name", "Math", "Science", "Total", "Status"]
ws.append(headers)

# Sample data
data = [
    ["AIK24CS001", "John", 85, 90, 175, "PASS"],
    ["AIK24CS002", "Jane", 45, 50, 95, "FAIL"],
    ["AIK24CS003", "Bob", 78, 82, 160, "PASS"],
]

for row in data:
    ws.append(row)

# Style the header row
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in ws[1]:  # Row 1 (header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Color-code the Status column
for row in range(2, ws.max_row + 1):  # Skip header
    status_cell = ws.cell(row=row, column=6)  # Column F (Status)
    
    if status_cell.value == "PASS":
        # Green background for PASS
        status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        status_cell.font = Font(bold=True, color="006100")
    else:
        # Red background for FAIL
        status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        status_cell.font = Font(bold=True, color="9C0006")

# Adjust column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 10

wb.save('styled_report.xlsx')
print("✓ Created 'styled_report.xlsx'")
print("  Open it - looks professional! 🎨")

print("\n" + "="*60)
print("LESSON 5: Working with Multiple Sheets")
print("-"*60)

wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# Create multiple sheets
summary_sheet = wb.create_sheet("Summary", 0)  # First sheet
ce_sheet = wb.create_sheet("Civil Engineering")
me_sheet = wb.create_sheet("Mechanical Engineering")
cs_sheet = wb.create_sheet("Computer Science")

# Add data to summary
summary_sheet['A1'] = "Department Summary"
summary_sheet['A3'] = "Civil Engineering"
summary_sheet['B3'] = "25 students"
summary_sheet['A4'] = "Mechanical"
summary_sheet['B4'] = "30 students"

# Add data to department sheets
ce_sheet['A1'] = "Civil Engineering Students"
ce_sheet.append(["Register", "Arrears"])
ce_sheet.append(["AIK24CE001", 2])

wb.save('multi_sheet.xlsx')
print("✓ Created 'multi_sheet.xlsx' with 4 sheets")
print("  Check all the tabs at the bottom!")
